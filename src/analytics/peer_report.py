import sqlite3
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DB_PATH = "db/nifty100.db"
PEER_FILE = "data/raw/peer_groups.xlsx"
OUTPUT_FILE = "output/peer_comparison.xlsx"


GREEN = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

YELLOW = PatternFill(
    start_color="FFEB9C",
    end_color="FFEB9C",
    fill_type="solid"
)

RED = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

GOLD = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid"
)


def load_data():

    conn = sqlite3.connect(DB_PATH)

    peer_percentiles = pd.read_sql(
        """
        SELECT *
        FROM peer_percentiles
        """,
        conn
    )

    conn.close()

    peer_groups = pd.read_excel(
        PEER_FILE
    )

    return peer_percentiles, peer_groups


def create_excel():

    peer_percentiles, peer_groups = load_data()

    peer_names = sorted(
        peer_percentiles["peer_group_name"].unique()
    )

    writer = pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    )

    for group in peer_names:

        group_df = peer_percentiles[
            peer_percentiles["peer_group_name"] == group
        ].copy()

        pivot = group_df.pivot_table(
            index="company_id",
            columns="metric",
            values="percentile_rank"
        )

        pivot.reset_index(inplace=True)

        medians = {}

        for col in pivot.columns:

            if col == "company_id":
                medians[col] = "Median"

            else:
                medians[col] = pivot[col].median()

        pivot.loc[len(pivot)] = medians

        pivot.to_excel(
            writer,
            sheet_name=group[:31],
            index=False
        )

    writer.close()

    format_workbook(peer_groups)

    print(
        f"Excel report generated -> {OUTPUT_FILE}"
    )


def format_workbook(peer_groups):

    wb = load_workbook(OUTPUT_FILE)

    benchmark_lookup = dict(
        zip(
            peer_groups["company_id"],
            peer_groups["is_benchmark"]
        )
    )

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        headers = [
            cell.value
            for cell in ws[1]
        ]

        percentile_cols = []

        for idx, col in enumerate(headers, start=1):

            if col != "company_id":

                percentile_cols.append(idx)

        for row in range(2, ws.max_row + 1):

            company = ws.cell(
                row=row,
                column=1
            ).value

            if company == "Median":

                continue

            if benchmark_lookup.get(company, False):

                for col in range(
                    1,
                    ws.max_column + 1
                ):
                    ws.cell(
                        row=row,
                        column=col
                    ).fill = GOLD

            for col in percentile_cols:

                cell = ws.cell(
                    row=row,
                    column=col
                )

                try:

                    value = float(cell.value)

                    if value >= 75:

                        cell.fill = GREEN

                    elif value <= 25:

                        cell.fill = RED

                    else:

                        cell.fill = YELLOW

                except:
                    pass

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    create_excel()
    