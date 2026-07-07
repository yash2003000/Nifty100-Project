import pandas as pd
from openpyxl.styles import PatternFill

from src.screener.presets import run_preset
PRESETS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch"
]
def export_screeners():

    with pd.ExcelWriter(
        "output/screener_output.xlsx",
        engine="openpyxl"
    ) as writer:

        for preset in PRESETS:

            df = run_preset(preset)

            df.to_excel(
                writer,
                sheet_name=preset[:31],
                index=False
            )
            worksheet = writer.sheets[preset[:31]]

            green_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
            )

        for row in range(2, len(df) + 2):
          worksheet.cell(
            row=row,
            column=1
          ).fill = green_fill

    print(
        "screener_output.xlsx generated"
    )
